import sys
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout, QHBoxLayout, QDesktopWidget, QFileDialog
from PyQt5.QtCore import Qt
import openpyxl

class SignInApp(QWidget):
    def __init__(self):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('签到应用')

        screen = QDesktopWidget().screenGeometry()
        window_size = self.geometry()
        self.setGeometry((screen.width() - window_size.width()) // 2,
                         (screen.height() - window_size.height()) // 2,
                         500, 300)
        
        self.choose_list_button = QPushButton('选择名单', self)
        self.label = QLabel('请输入要签到的姓名:', self)
        self.name_input = QLineEdit(self)
        self.sign_in_button = QPushButton('签到', self)
        self.result_label = QLabel('', self)


        layout = QVBoxLayout()
        layout.addWidget(self.choose_list_button)
        layout.addWidget(self.label)
        layout.addWidget(self.name_input)

        layout.addWidget(self.sign_in_button)
        layout.addWidget(self.result_label)

        self.sign_in_button.clicked.connect(self.perform_sign_in)
        self.name_input.returnPressed.connect(self.perform_sign_in)
        self.choose_list_button.clicked.connect(self.choose_sign_in_list)

        h_layout = QHBoxLayout()
        h_layout.addStretch(1)
        h_layout.addWidget(self.sign_in_button)
        h_layout.addStretch(1)

        layout.addLayout(h_layout)

        self.setLayout(layout)

    def perform_sign_in(self):
        name_to_check = self.name_input.text()
        result = self.sign_in(self.selected_file, name_to_check)
        self.result_label.setText(result)

        # Clear the input field after sign-in
        self.name_input.clear()

    def sign_in(self, file_path, name):
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] == name:
                    sheet.cell(row=idx, column=2, value='已签到')
                    workbook.save(file_path)
                    return f"{name} 已签到"

            return f"未找到姓名：{name}"

        except FileNotFoundError:
            return f"文件未找到：{file_path}"
        except Exception as e:
            return f"读取文件时出错：{e}"

    def choose_sign_in_list(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("Excel files (*.xlsx)")
        file_dialog.setOptions(options)

        if file_dialog.exec_() == QFileDialog.Accepted:
            self.selected_file = file_dialog.selectedFiles()[0]
            self.result_label.setText(f"已选择名单：{self.selected_file}")

def main():
    app = QApplication(sys.argv)
    window = SignInApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
